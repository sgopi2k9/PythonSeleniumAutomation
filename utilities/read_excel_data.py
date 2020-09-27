import openpyxl

def readExcelData(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    my_data = []
    t=[]
    maxRow = ws.max_row
    maxColumn = ws.max_column

    for i in range(2,maxRow+1):
        for j in range(1,maxColumn):
            t.append(ws.cell(row=i,column=j).value)
        my_data.append(list(t))
        t.clear()
    return my_data

def writeExcelData(filepath,row_num,column_num, result_message):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    ws.cell(row=row_num,column=column_num).value = result_message
    wb.save(filepath)


"""
def writeExcelData(filepath, result):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.get_sheet_by_name("Sheet 1")
    ws.
"""
#readExcelData()
